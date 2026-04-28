import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
import io

# --- CONFIGURAZIONE FILE ---
FILE_MODELLO = "calcolatore_26.xlsm"
FILE_CARATTERISTICHE = "database_caratteristiche.xlsx" 
FOGLIO_INPUT = "INSER. DATI"

# --- FUNZIONE CARICAMENTO OPZIONI ---
def carica_opzioni(file_name):
    if os.path.exists(file_name):
        for enc in ['utf-8-sig', 'latin1', 'cp1252']:
            try:
                df = pd.read_csv(file_name, sep=None, engine='python', encoding=enc, on_bad_lines='skip')
                return df.iloc[:, 0].dropna().astype(str).str.strip().unique().tolist()
            except: continue
    return []

lista_localita = carica_opzioni("database_localita.csv")
lista_fanali = carica_opzioni("database_fanali.csv")
lista_colori = carica_opzioni("database_colori.csv")
lista_portate = ["", "1", "2 (-)", "2", "3 (-)", "3", "4 (-)", "4", "5 (-)", "5", "6 (-)", "6", "7 (-)", "7", "8 (-)", "8", "9 (-)", "9"]

st.set_page_config(page_title="Calcolatore Solare", layout="centered")
st.markdown('### 📋 CALCOLATORE SOLARE by Emmekappa')

# --- RICERCA CARATTERISTICA (0-9 senza 8) ---
lampi_validi = [0, 1, 2, 3, 4, 5, 6, 7, 9]
n_lampi = st.selectbox("Numero lampi per caratteristica", options=lampi_validi)

opzioni_mostrate = [] 
dato_scelto = ""

if n_lampi == 0:
    dato_scelto = "luce fissa"
    st.info("Selezionato: luce fissa")
else:
    if os.path.exists(FILE_CARATTERISTICHE):
        try:
            df_car = pd.read_excel(FILE_CARATTERISTICHE, dtype=str)
            filtro = df_car[df_car.iloc[:, 0].str.strip() == str(n_lampi)]
            for _, row in filtro.iterrows():
                for cella in row.values:
                    if pd.notna(cella):
                        v = str(cella).strip()
                        if "sec." in v or "=" in v or "-" in v:
                            opzioni_mostrate.append(v)
                            break
        except: pass
    dato_scelto = st.selectbox("Seleziona la caratteristica", options=opzioni_mostrate if opzioni_mostrate else [""])

# --- MODULO ---
with st.form("form_inserimento"):
    col1, col2 = st.columns(2)
    with col1:
        cliente = st.text_input("Cliente")
        commessa = st.text_input("COMM. N°")
        loc_sel = st.selectbox("Località", [""] + lista_localita)
    with col2:
        fan_sel = st.selectbox("Tipo di Fanale", [""] + lista_fanali)
        col_sel = st.selectbox("Colore della luce", [""] + lista_colori)
        portata_sel = st.selectbox("Portata", lista_portate)
        gps_sel = st.radio("GPS", ["COMPRESO", "NON COMPRESO"], index=None, horizontal=True)
    submit = st.form_submit_button("GENERA FILE AGGIORNATO 💾")

if submit:
    if not all([cliente, commessa, fan_sel, loc_sel, col_sel, portata_sel, gps_sel]):
        st.error("⚠️ Compila tutti i campi!")
    else:
        try:
            book = load_workbook(FILE_MODELLO, keep_vba=True)
            ws = book[FOGLIO_INPUT]
            
            # Scrittura in colonna S (appoggio)
            ws["S4"], ws["S5"], ws["S6"], ws["S8"] = fan_sel, loc_sel, col_sel, dato_scelto
            ws["S12"], ws["S16"], ws["S26"], ws["S27"] = portata_sel, gps_sel, commessa, cliente
            
            output = io.BytesIO()
            book.save(output)
            nome_file_personalizzato = f"Comm. {commessa}_{cliente}.xlsm"
            
            st.success(f"✅ Dati pronti per {cliente}!")
            st.download_button(
                label=f"📥 Scarica: {nome_file_personalizzato}",
                data=output.getvalue(),
                file_name=nome_file_personalizzato,
                mime="application/vnd.ms-excel.sheet.macroEnabled.12"
            )
            book.close()
        except Exception as e:
            st.error(f"Errore: {e}")
